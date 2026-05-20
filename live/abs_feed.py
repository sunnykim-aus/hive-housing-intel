"""
Live feed from ABS (Australian Bureau of Statistics).
Downloads the latest Building Approvals and Property Price data.
Data is cached locally and refreshed monthly.
"""
import io
import json
import re
import httpx
import openpyxl
from datetime import datetime, date
from pathlib import Path
from bs4 import BeautifulSoup

from config import DATA_DIR

CACHE_DIR = DATA_DIR / "live_cache"
CACHE_DIR.mkdir(parents=True, exist_ok=True)

ABS_BASE = "https://www.abs.gov.au"
BA_LISTING = "/statistics/industry/building-and-construction/building-approvals-australia/latest-release"
RPPI_LISTING = "/statistics/economy/price-indexes-and-inflation/residential-property-price-indexes-eight-capital-cities/latest-release"

HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"}


def _cache_path(name):
    return CACHE_DIR / f"{name}.json"


def _is_stale(path, days=30):
    if not path.exists():
        return True
    age = (datetime.now().timestamp() - path.stat().st_mtime) / 86400
    return age > days


def _find_excel_url(listing_url, table_num="8731001"):
    """Find the latest Excel download URL from an ABS release page."""
    try:
        r = httpx.get(ABS_BASE + listing_url, headers=HEADERS, timeout=20, follow_redirects=True)
        soup = BeautifulSoup(r.text, "lxml")
        for a in soup.select("a[href]"):
            href = a.get("href", "")
            if table_num in href and href.endswith(".xlsx"):
                return ABS_BASE + href
    except Exception:
        pass
    return None


def _download_excel(url):
    try:
        r = httpx.get(url, headers=HEADERS, timeout=60, follow_redirects=True)
        if r.status_code == 200:
            return io.BytesIO(r.content)
    except Exception:
        pass
    return None


def fetch_building_approvals(force=False):
    """
    Returns monthly total dwelling approvals for Australia.
    Columns: date, total_aus, houses_aus, other_aus
    Last 5 years of data.
    """
    cache = _cache_path("building_approvals")
    if not force and not _is_stale(cache, days=30):
        return json.loads(cache.read_text())

    # File 8731009 has all states + Australia total
    url = _find_excel_url(BA_LISTING, "8731009")
    if not url:
        year_month = datetime.now().strftime("%b-%Y").lower()
        url = f"{ABS_BASE}/statistics/industry/building-and-construction/building-approvals-australia/{year_month}/8731009.xlsx"

    buf = _download_excel(url)
    if not buf:
        return _load_cache_or_empty(cache)

    try:
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        ws = wb["Data1"]
        rows = list(ws.iter_rows(values_only=True))
        headers_row = rows[0]

        # Find Australia total columns (confirmed: col 27=total, col 9=houses, col 18=other)
        total_aus_col = houses_aus_col = other_aus_col = None
        for i, h in enumerate(headers_row):
            if not h:
                continue
            h_str = str(h)
            if "Australia" in h_str and "Total (Type of Building)" in h_str:
                total_aus_col = i
            elif "Australia" in h_str and " Houses ;" in h_str:
                houses_aus_col = i
            elif "Australia" in h_str and "Dwellings excluding houses" in h_str:
                other_aus_col = i

        records = []
        cutoff = datetime(2015, 1, 1)
        for row in rows[10:]:
            dt = row[0]
            if not isinstance(dt, datetime) or dt < cutoff:
                continue
            records.append({
                "date": dt.strftime("%Y-%m-%d"),
                "total_aus": row[total_aus_col] if total_aus_col is not None else None,
                "houses_aus": row[houses_aus_col] if houses_aus_col is not None else None,
                "other_aus": row[other_aus_col] if other_aus_col is not None else None,
            })

        result = {
            "source": "ABS Building Approvals (Cat. 8731.0)",
            "fetched": date.today().isoformat(),
            "latest_period": records[-1]["date"] if records else None,
            "records": records[-60:],  # last 5 years
        }
        cache.write_text(json.dumps(result))
        return result

    except Exception as e:
        print(f"  [warn] Building approvals parse failed: {e}")
        return _load_cache_or_empty(cache)


def fetch_state_approvals(force=False):
    """
    Returns latest 12 months of approvals by state.
    """
    cache = _cache_path("state_approvals")
    if not force and not _is_stale(cache, days=30):
        return json.loads(cache.read_text())

    url = _find_excel_url(BA_LISTING, "8731002")
    if not url:
        year_month = datetime.now().strftime("%b-%Y").lower()
        url = f"{ABS_BASE}/statistics/industry/building-and-construction/building-approvals-australia/{year_month}/8731002.xlsx"

    buf = _download_excel(url)
    if not buf:
        return _load_cache_or_empty(cache)

    try:
        wb = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        ws = wb["Data1"]
        rows = list(ws.iter_rows(values_only=True))

        headers_row = rows[0]
        states = ["New South Wales", "Victoria", "Queensland", "Western Australia",
                  "South Australia", "Tasmania", "Australian Capital Territory", "Northern Territory"]

        state_cols = {}
        for i, h in enumerate(headers_row):
            if not h:
                continue
            h_str = str(h)
            for state in states:
                if state in h_str and "Total (Type of Building)" in h_str and "Total Sectors" in h_str:
                    state_cols[state[:3]] = i  # NSW, VIC, etc.

        cutoff = datetime(2023, 1, 1)
        records = []
        for row in rows[10:]:
            dt = row[0]
            if not isinstance(dt, datetime) or dt < cutoff:
                continue
            rec = {"date": dt.strftime("%Y-%m-%d")}
            for abbr, col in state_cols.items():
                rec[abbr] = row[col]
            records.append(rec)

        result = {
            "source": "ABS Building Approvals by State (Cat. 8731.0)",
            "fetched": date.today().isoformat(),
            "state_cols": list(state_cols.keys()),
            "records": records[-24:],
        }
        cache.write_text(json.dumps(result))
        return result

    except Exception as e:
        print(f"  [warn] State approvals parse failed: {e}")
        return _load_cache_or_empty(cache)


def fetch_housing_indicators():
    """
    Returns a summary dict of key current housing indicators.
    Combines data from building approvals fetch.
    """
    ba = fetch_building_approvals()
    records = ba.get("records", [])
    if not records:
        return {}

    latest = records[-1]
    prev_year = next((r for r in reversed(records[:-1]) if r["date"][:4] != latest["date"][:4]), None)
    three_months = records[-3:]
    avg_3m = sum(r["total_aus"] or 0 for r in three_months) / 3 if three_months else 0
    annual_run_rate = avg_3m * 12

    yoy_change = None
    if prev_year and prev_year.get("total_aus") and latest.get("total_aus"):
        yoy_change = ((latest["total_aus"] - prev_year["total_aus"]) / prev_year["total_aus"]) * 100

    national_accord_target = 240000  # dwellings per year (1.2M over 5 years)
    gap_to_target = annual_run_rate - national_accord_target if annual_run_rate else None

    return {
        "latest_month": latest["date"],
        "latest_total": latest.get("total_aus"),
        "latest_houses": latest.get("houses_aus"),
        "latest_other": latest.get("other_aus"),
        "annual_run_rate": round(annual_run_rate),
        "yoy_change_pct": round(yoy_change, 1) if yoy_change else None,
        "national_accord_target": national_accord_target,
        "gap_to_accord_target": round(gap_to_target) if gap_to_target else None,
        "source": ba.get("source"),
        "fetched": ba.get("fetched"),
    }


def _load_cache_or_empty(cache_path):
    if cache_path.exists():
        return json.loads(cache_path.read_text())
    return {"records": [], "error": "Data unavailable"}
